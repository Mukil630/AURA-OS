import os
import pytest
from brain.codeact_cloud_runner import CodeActCloudRunner


def test_codeact_runner_simple_execution():
    runner = CodeActCloudRunner()
    code = "import math\nprint(f'PI={math.pi:.4f}')"
    res = runner.execute_script(code, task_name="test_pi")
    assert res.status == "SUCCESS"
    assert "PI=3.1416" in res.stdout


def test_codeact_runner_excel_generation():
    runner = CodeActCloudRunner()
    excel_path = os.path.join(runner.scratch_dir, "test_output.xlsx")
    code = f"""
import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 'Company'
ws['B1'] = 'Status'
ws.append(['Karur Textiles', 'Verified'])
wb.save(r'{excel_path}')
print('EXCEL_SAVED')
"""
    res = runner.execute_script(code, task_name="test_excel")
    assert res.status == "SUCCESS"
    assert "EXCEL_SAVED" in res.stdout
    assert os.path.exists(excel_path)
    if os.path.exists(excel_path):
        os.remove(excel_path)
